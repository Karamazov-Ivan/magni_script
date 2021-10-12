import os
import settings
import openpyxl

# Берём директорию из настроек
directory = settings.NEW_FILE_PATH

# Сохраняем имена всех файлов в директории
files = os.listdir(directory)

new_folder = settings.PATH + "/Один_файл"

# Разкомментировать, если надо включить проверку существования папки
# if "Один_файл" in files:
#     inp = input("Папка 'Один_файл' уже существует. Перезаписать?\ny/n\n")
#     if inp == "y":
#         pass
#     else:
#         exit()

# if "Один_файл" not in files:
#     # Создаём новую папку
#     os.mkdir(new_folder)   

# Создаём новую книгу для объединения всех
wb_two = openpyxl.Workbook()
sheet_two = wb_two.active

# Создаём книгу для отчёта об ошибках и делаем шапку
wb_error = openpyxl.Workbook()
sheet_error = wb_error.active
error_hat = ['Название файла', 'Содержание ошибочной ячейки', 'Как должно быть', 'Ошибка']
accum = -1
for colum in range(1, len(error_hat) + 1):
    accum += 1
    sheet_error.cell(row=1, column=colum).value = error_hat[accum]

# Открываем шаблон
wb_reference =  openpyxl.reader.excel.load_workbook(filename=settings.REFERENCE_FILE)
wb_reference.active = 0

wb_reference_sheet = wb_reference.active

# Переменная-накопитель для записи в строки
world_ro = 0
error_row = 1

# Индикатор наличия ошибок
error_ind = 0

for file_name in files:
    if file_name.endswith(".xlsx"):
        # Открываем файл
        wb_one = openpyxl.reader.excel.load_workbook(filename=directory + '/' + file_name)

        # Назначаем активным первый Лист
        wb_one.active = 0

        # Сохраняем в переменную активный лист
        sheet = wb_one.active

        # Проверка на соответствие шаблону по ко-ву строк
        if sheet.max_row != wb_reference_sheet.max_row:
            error_row += 1
            sheet_error.cell(row=error_row, column=1).value = file_name
            sheet_error.cell(row=error_row, column=2).value = 'н/д'
            sheet_error.cell(row=error_row, column=3).value = 'н/д'
            sheet_error.cell(row=error_row, column=4).value = 'Несоответствует кол-во строк'
            print(f"Файл '{file_name}', ошибка: 'Несоответствует кол-во строк'")
            error_ind = 1
        
        # Проверка на соответствие шаблону по совпадению ячеек
        for ro in range(2, wb_reference_sheet.max_row + 1):
            for co in range(1, wb_reference_sheet.max_column - 1):
                if sheet.cell(row=ro, column=co).value != wb_reference_sheet.cell(row=ro, column=co).value:
                    error_row += 1
                    sheet_error.cell(row=error_row, column=1).value = file_name
                    sheet_error.cell(row=error_row, column=2).value = sheet.cell(row=ro, column=co).value
                    sheet_error.cell(row=error_row, column=3).value = wb_reference_sheet.cell(row=ro, column=co).value
                    sheet_error.cell(row=error_row, column=4).value = 'Несоответствие ячеек'
                    print(f"Файл '{file_name}', ошибка: 'Несоответствие ячеек'")
                    error_ind = 1



        for ro in range(2, sheet.max_row + 1):
            world_ro += 1
            for co in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=ro, column=co).value
                sheet_two.cell(row=world_ro, column=co).value = cell_value
                               
        print(f"Файл '{file_name[:-5]}' записан")
    
# Сохраняем файл
wb_two.save(str(new_folder) + '/' + 'Совершенно_новый_файл.xlsx')
wb_error.save(str(new_folder) + '/' + 'Отчёт_об_ошибках.xlsx')

if error_ind == 0:
    print("Всё сработало, расходимся...")
else:
    print("Есть ошибки, проверь отчёт!!!")
