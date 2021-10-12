import os
import settings
import openpyxl

# Берём директорию из настроек
directory = settings.NEW_FILE_PATH

# Сохраняем имена всех файлов в директории
files = os.listdir(directory) 

new_folder = settings.PATH + "/Один_файл"

# Разкомментировать, если надо включить проверку существования папки
# if "Один_файл" in files:
#     inp = input("Папка 'Один_файл' уже существует. Перезаписать?\ny/n\n")
#     if inp == "y":
#         pass
#     else:
#         exit()

# if "Один_файл" not in files:
#     # Создаём новую папку
#     os.mkdir(new_folder)   

world_ro = 0

wb_two = openpyxl.Workbook()

sheet_two = wb_two.active

for file_name in files:
    if file_name.endswith(".xlsx"):
        # Открываем файл
        wb_one = openpyxl.reader.excel.load_workbook(filename=directory + '/' + file_name)

        # Назначаем активным первый Лист
        wb_one.active = 0

        # Сохраняем в переменную активный лист
        sheet = wb_one.active
        
        for ro in range(2, sheet.max_row + 1):
            world_ro += 1
            for co in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=ro, column=co).value
                sheet_two.cell(row=world_ro, column=co).value = cell_value
                               
        print(f"Файл '{file_name[:-5]}' записан")
    
# Сохраняем файл
wb_two.save(str(new_folder) + '/' + 'Совершенно_новый_файл.xlsx')

print("Всё сработало, расходимся...")
